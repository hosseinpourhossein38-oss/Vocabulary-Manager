from __future__ import annotations

import csv
import io
import json
import os
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Any, Dict, List, Tuple
from urllib.parse import urljoin, urlparse

from dotenv import load_dotenv
load_dotenv(override=True)

from flask import (
    Flask, Response, flash, redirect, render_template,
    request, session, url_for, jsonify
)
from flask_login import (
    LoginManager, UserMixin, current_user,
    login_required, login_user, logout_user,
)
from werkzeug.security import check_password_hash
from werkzeug.utils import secure_filename

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Font, Side
from openpyxl.utils import get_column_letter

from sqlalchemy import func
from sqlalchemy.orm import Session

# --- DB & Models ---
from db import engine, SessionLocal, init_db
from models import Vocabulary, Goal, TimeEntry
init_db()
# ------------------------------------------------------------------------------
# App & Auth setup
# ------------------------------------------------------------------------------
app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "dev-secret-change-me")

# initialize DB (create tables if missing)
init_db()

ADMIN_EMAIL = (os.environ.get("ADMIN_EMAIL") or "").strip()
ADMIN_HASH = (os.environ.get("ADMIN_HASH") or "").strip()

login_manager = LoginManager(app)
login_manager.login_view = "login"
login_manager.login_message_category = "warning"


class AdminUser(UserMixin):
    def __init__(self, email: str) -> None:
        self._email = email

    def get_id(self) -> str:
        return self._email


@login_manager.user_loader
def load_user(user_id: str) -> AdminUser | None:
    if ADMIN_EMAIL and user_id == ADMIN_EMAIL:
        return AdminUser(user_id)
    return None


# ------------------------------------------------------------------------------
# Constants / helpers
# ------------------------------------------------------------------------------
MAX_FILE_SIZE_MB = 2
app.config["MAX_CONTENT_LENGTH"] = MAX_FILE_SIZE_MB * 1024 * 1024

# همان فیلدهایی که قبلاً در اکسپورت‌ها داشتی
EXPORT_FIELDS = [
    "word",
    "sentence",
    "synonym",
    "native_meaning",
    "type",
    "base_word",
    "created_at",
]
IMPORT_FIELDS = EXPORT_FIELDS
DEDUP_FIELDS = ("word", "sentence", "synonym", "type", "base_word", "native_meaning")
ALLOWED_IMPORT_EXTENSIONS = {".csv", ".json", ".xlsx"}
PENDING_IMPORT_SESSION_KEY = "pending_import_rows"
PENDING_IMPORT_META_SESSION_KEY = "pending_import_meta"


def export_timestamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%SZ")


def _str(v: Any) -> str:
    return "" if v is None else str(v).strip()


def _norm_key(k: Any) -> str:
    return "" if k is None else str(k).strip().lower().replace(" ", "_")


def _serialize_vocab(v: Vocabulary) -> Dict[str, Any]:
    """ORM -> dict with same keys the templates expect (type instead of pos)."""
    return {
        "id": str(v.id),
        "word": v.word or "",
        "sentence": v.sentence or "",
        "synonym": v.synonym or "",
        "native_meaning": v.native_meaning or "",
        "type": v.pos or "",          # map pos -> type
        "base_word": v.base_word or "",
        "created_at": (v.created_at or datetime.now(timezone.utc)).isoformat(),
    }


def _dedup_key_dict(d: Dict[str, Any]) -> Tuple[str, str, str, str, str, str]:
    return tuple(_str(d.get(k)) for k in DEDUP_FIELDS)


# ------------------------------------------------------------------------------
# Healthz
# ------------------------------------------------------------------------------
@app.route("/healthz", methods=["GET"])
def healthz() -> Response:
    return Response("OK", status=200, mimetype="text/plain")


# ------------------------------------------------------------------------------
# Auth views
# ------------------------------------------------------------------------------
def _is_safe_redirect(target: str) -> bool:
    if not target:
        return False
    host_url = urlparse(request.host_url)
    redirect_url = urlparse(urljoin(request.host_url, target))
    return redirect_url.scheme in ("http", "https") and host_url.netloc == redirect_url.netloc


@app.route("/login", methods=["GET"])
def login() -> Response | str:
    next_url = request.args.get("next")
    if next_url and not _is_safe_redirect(next_url):
        next_url = None
    if current_user.is_authenticated:
        return redirect(next_url or url_for("index"))
    return render_template("login.html", next_url=next_url)


@app.route("/login", methods=["POST"])
def login_submit() -> Response:
    email = (request.form.get("email") or "").strip()
    password = request.form.get("password") or ""
    next_url = request.form.get("next")
    if next_url and not _is_safe_redirect(next_url):
        next_url = None

    if not ADMIN_EMAIL or not ADMIN_HASH:
        flash("Authentication is not configured.", "error")
        return redirect(url_for("login", next=next_url) if next_url else url_for("login"))

    try:
        ok = check_password_hash(ADMIN_HASH, password)
    except ValueError:
        flash("Authentication is misconfigured.", "error")
        return redirect(url_for("login", next=next_url) if next_url else url_for("login"))

    if email == ADMIN_EMAIL and ok:
        login_user(AdminUser(ADMIN_EMAIL))
        flash("Logged in successfully.", "success")
        return redirect(next_url or url_for("index"))

    flash("Invalid email or password.", "error")
    return redirect(url_for("login", next=next_url) if next_url else url_for("login"))


@app.route("/logout", methods=["POST"])
@login_required
def logout() -> Response:
    logout_user()
    flash("Logged out.", "success")
    return redirect(url_for("index"))


# ------------------------------------------------------------------------------
# Global auth guard
# ------------------------------------------------------------------------------
@app.before_request
def require_auth():
    if request.endpoint in ["login", "login_submit", "logout", "healthz", "static"]:
        return None
    if not current_user.is_authenticated:
        return redirect(url_for("login", next=request.url))


# ------------------------------------------------------------------------------
# Dashboard
# ------------------------------------------------------------------------------
@app.route("/dashboard", methods=["GET"])
@login_required
def dashboard() -> str:
    db: Session = SessionLocal()
    try:
        # vocab summary
        last5 = db.query(Vocabulary).order_by(Vocabulary.created_at.desc()).limit(5).all()
        recent_vocab = [_serialize_vocab(v) for v in last5]
        vocab_total = db.query(func.count(Vocabulary.id)).scalar() or 0

        # goals + time
        active_goals = db.query(Goal).filter(Goal.is_active.is_(True)).all()

        today = datetime.now().date()
        week_start = today - timedelta(days=today.weekday())
        month_start = today.replace(day=1)

        today_time = db.query(func.sum(TimeEntry.minutes)).filter(
            func.date(TimeEntry.started_at) == today
        ).scalar() or 0

        week_time = db.query(func.sum(TimeEntry.minutes)).filter(
            func.date(TimeEntry.started_at) >= week_start
        ).scalar() or 0

        month_time = db.query(func.sum(TimeEntry.minutes)).filter(
            func.date(TimeEntry.started_at) >= month_start
        ).scalar() or 0

        return render_template(
            "dashboard.html",
            vocab_total=vocab_total,
            recent_vocab=recent_vocab,
            active_goals=active_goals,
            today_time=today_time,
            week_time=week_time,
            month_time=month_time,
        )
    finally:
        db.close()


# ------------------------------------------------------------------------------
# Goals CRUD (بدون تغییرات جدی)
# ------------------------------------------------------------------------------
@app.route("/goals", methods=["GET"])
@login_required
def goals_list() -> str:
    db = SessionLocal()
    try:
        goals = db.query(Goal).order_by(Goal.created_at.desc()).all()
        return render_template("goals.html", goals=goals)
    finally:
        db.close()


@app.route("/goals", methods=["POST"])
@login_required
def create_goal() -> Response:
    title = (request.form.get("title") or "").strip()
    category = (request.form.get("category") or "").strip()
    notes = (request.form.get("notes") or "").strip()

    if not title:
        flash("Title is required.", "error")
        return redirect(url_for("goals_list"))

    db = SessionLocal()
    try:
        goal = Goal(title=title, category=category, notes=notes)
        db.add(goal)
        db.commit()
        flash("Goal created successfully!", "success")
        return redirect(url_for("goals_list"))
    finally:
        db.close()


@app.route("/goals/<int:goal_id>/edit", methods=["POST"])
@login_required
def update_goal(goal_id: int) -> Response:
    title = (request.form.get("title") or "").strip()
    category = (request.form.get("category") or "").strip()
    notes = (request.form.get("notes") or "").strip()
    is_active = request.form.get("is_active") == "on"

    if not title:
        flash("Title is required.", "error")
        return redirect(url_for("goals_list"))

    db = SessionLocal()
    try:
        goal = db.query(Goal).filter(Goal.id == goal_id).first()
        if not goal:
            flash("Goal not found.", "error")
            return redirect(url_for("goals_list"))

        goal.title = title
        goal.category = category
        goal.notes = notes
        goal.is_active = is_active
        db.commit()
        flash("Goal updated successfully!", "success")
        return redirect(url_for("goals_list"))
    finally:
        db.close()


@app.route("/goals/<int:goal_id>/delete", methods=["POST"])
@login_required
def delete_goal(goal_id: int) -> Response:
    db = SessionLocal()
    try:
        goal = db.query(Goal).filter(Goal.id == goal_id).first()
        if not goal:
            flash("Goal not found.", "error")
            return redirect(url_for("goals_list"))

        db.delete(goal)
        db.commit()
        flash("Goal deleted successfully!", "success")
        return redirect(url_for("goals_list"))
    finally:
        db.close()


@app.route("/goals/<int:goal_id>/log", methods=["GET"])
@login_required
def log_time_form(goal_id: int) -> str:
    db = SessionLocal()
    try:
        goal = db.query(Goal).filter(Goal.id == goal_id).first()
        if not goal:
            flash("Goal not found.", "error")
            return redirect(url_for("goals_list"))

        recent_entries = (
            db.query(TimeEntry)
            .filter(TimeEntry.goal_id == goal_id)
            .order_by(TimeEntry.started_at.desc())
            .limit(10)
            .all()
        )
        return render_template("goal_detail.html", goal=goal, recent_entries=recent_entries)
    finally:
        db.close()


@app.route("/goals/<int:goal_id>/log", methods=["POST"])
@login_required
def log_time(goal_id: int) -> Response:
    started_at_str = (request.form.get("started_at") or "").strip()
    minutes = (request.form.get("minutes") or "").strip()
    note = (request.form.get("note") or "").strip()

    if not started_at_str or not minutes:
        flash("Started at and minutes are required.", "error")
        return redirect(url_for("log_time_form", goal_id=goal_id))

    try:
        started_at = datetime.fromisoformat(started_at_str.replace("Z", "+00:00"))
        minutes_int = int(minutes)
    except (ValueError, TypeError):
        flash("Invalid date or minutes format.", "error")
        return redirect(url_for("log_time_form", goal_id=goal_id))

    db = SessionLocal()
    try:
        goal = db.query(Goal).filter(Goal.id == goal_id).first()
        if not goal:
            flash("Goal not found.", "error")
            return redirect(url_for("goals_list"))

        entry = TimeEntry(goal_id=goal_id, started_at=started_at, minutes=minutes_int, note=note)
        db.add(entry)
        db.commit()
        flash("Time logged successfully!", "success")
        return redirect(url_for("log_time_form", goal_id=goal_id))
    finally:
        db.close()


# ------------------------------------------------------------------------------
# Vocabulary views (DB-based)
# ------------------------------------------------------------------------------
def _get_vocab_list() -> List[Dict[str, Any]]:
    db = SessionLocal()
    try:
        rows = db.query(Vocabulary).order_by(Vocabulary.created_at.desc()).all()
        return [_serialize_vocab(v) for v in rows]
    finally:
        db.close()


@app.route("/", methods=["GET"])
@login_required
def index() -> str:
    items = _get_vocab_list()
    return render_template("index.html", items=items, total=len(items), edit_entry_id=None)


@app.route("/vocabulary", methods=["GET"])
@login_required
def vocabulary() -> str:
    items = _get_vocab_list()
    return render_template("index.html", items=items, total=len(items), edit_entry_id=None)


@app.route("/add", methods=["POST"])
@login_required
def add_word() -> Response:
    word = _str(request.form.get("word"))
    sentence = _str(request.form.get("sentence"))
    synonym = _str(request.form.get("synonym"))
    type_ = _str(request.form.get("type"))       # => pos
    base_word = _str(request.form.get("base_word"))
    native_meaning = _str(request.form.get("native_meaning"))

    if not word:
        flash("Field 'word' is required.", "error")
        return redirect(url_for("vocabulary"))

    db = SessionLocal()
    try:
        exists = (
            db.query(Vocabulary.id)
            .filter(
                Vocabulary.word == word,
                Vocabulary.sentence == sentence,
                Vocabulary.synonym == synonym,
                Vocabulary.pos == type_,
                Vocabulary.base_word == base_word,
                Vocabulary.native_meaning == native_meaning,
            )
            .first()
        )
        if exists:
            flash("This exact entry already exists.", "warning")
            return redirect(url_for("vocabulary"))

        new_word = Vocabulary(
            word=word,
            sentence=sentence,
            synonym=synonym,
            pos=type_,
            base_word=base_word,
            native_meaning=native_meaning,
            created_at=datetime.now(timezone.utc),
        )
        db.add(new_word)
        db.commit()
        flash("Word added successfully!", "success")
    except Exception as e:
        db.rollback()
        flash(f"Error while saving: {e}", "error")
    finally:
        db.close()

    return redirect(url_for("vocabulary"))


@app.route("/edit/<entry_id>", methods=["GET"])
@login_required
def edit_entry_form(entry_id: str) -> Response | str:
    items = _get_vocab_list()
    if not any(i["id"] == entry_id for i in items):
        flash("Entry not found.", "error")
        return redirect(url_for("vocabulary"))
    return render_template("index.html", items=items, total=len(items), edit_entry_id=entry_id)


@app.route("/edit/<entry_id>", methods=["POST"])
@login_required
def edit_entry(entry_id: str) -> Response:
    word = _str(request.form.get("word"))
    sentence = _str(request.form.get("sentence"))
    synonym = _str(request.form.get("synonym"))
    type_ = _str(request.form.get("type"))
    base_word = _str(request.form.get("base_word"))
    native_meaning = _str(request.form.get("native_meaning"))

    if not word:
        flash("Field 'word' is required.", "error")
        return redirect(url_for("edit_entry_form", entry_id=entry_id))

    db = SessionLocal()
    try:
        item = db.query(Vocabulary).filter(Vocabulary.id == entry_id).first()
        if not item:
            flash("Entry not found.", "error")
            return redirect(url_for("vocabulary"))

        dup = (
            db.query(Vocabulary.id)
            .filter(
                Vocabulary.id != entry_id,
                Vocabulary.word == word,
                Vocabulary.sentence == sentence,
                Vocabulary.synonym == synonym,
                Vocabulary.pos == type_,
                Vocabulary.base_word == base_word,
                Vocabulary.native_meaning == native_meaning,
            )
            .first()
        )
        if dup:
            flash("This exact entry already exists.", "warning")
            return redirect(url_for("edit_entry_form", entry_id=entry_id))

        item.word = word
        item.sentence = sentence
        item.synonym = synonym
        item.pos = type_
        item.base_word = base_word
        item.native_meaning = native_meaning
        db.commit()
        flash("Entry updated.", "success")
    except Exception as e:
        db.rollback()
        flash(f"Error: {e}", "error")
    finally:
        db.close()

    return redirect(url_for("vocabulary"))


@app.route("/delete/<entry_id>", methods=["POST"])
@login_required
def delete_entry(entry_id: str) -> Response:
    db = SessionLocal()
    try:
        item = db.query(Vocabulary).filter(Vocabulary.id == entry_id).first()
        if not item:
            flash("Entry not found.", "error")
            return redirect(url_for("vocabulary"))

        db.delete(item)
        db.commit()
        flash("Entry deleted.", "success")
    except Exception as e:
        db.rollback()
        flash(f"Error: {e}", "error")
    finally:
        db.close()

    return redirect(url_for("vocabulary"))


# ------------------------------------------------------------------------------
# Import helpers (parsers)
# ------------------------------------------------------------------------------
def _parse_csv(data: bytes) -> List[Dict[str, str]]:
    try:
        txt = data.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ValueError("CSV file must be UTF-8 encoded.") from exc
    reader = csv.DictReader(io.StringIO(txt))
    if reader.fieldnames is None:
        raise ValueError("CSV file is missing a header row.")
    rows: List[Dict[str, str]] = []
    for raw in reader:
        prepared = {_norm_key(k): raw[k] for k in raw.keys()}
        entry = {f: _str(prepared.get(f, "")) for f in IMPORT_FIELDS}
        if not any(entry[f] for f in DEDUP_FIELDS):
            continue
        rows.append(entry)
    return rows


def _parse_json(data: bytes) -> List[Dict[str, str]]:
    try:
        txt = data.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ValueError("JSON file must be UTF-8 encoded.") from exc
    try:
        payload: Any = json.loads(txt)
    except json.JSONDecodeError as exc:
        raise ValueError("JSON file is not valid JSON.") from exc

    if isinstance(payload, dict):
        if "items" in payload and isinstance(payload["items"], list):
            payload = payload["items"]
        else:
            payload = [payload]
    if not isinstance(payload, list):
        raise ValueError("JSON file must contain a list of entries.")

    rows: List[Dict[str, str]] = []
    for item in payload:
        if not isinstance(item, dict):
            continue
        prepared = {_norm_key(k): v for k, v in item.items()}
        entry = {f: _str(prepared.get(f, "")) for f in IMPORT_FIELDS}
        if not any(entry[f] for f in DEDUP_FIELDS):
            continue
        rows.append(entry)
    return rows


def _parse_xlsx(data: bytes) -> List[Dict[str, str]]:
    buf = io.BytesIO(data)
    wb = load_workbook(buf, read_only=True, data_only=True)
    try:
        sh = wb.active
        header = next(sh.iter_rows(values_only=True), None)
        if header is None:
            return []
        headers = [_norm_key(h) for h in header]
        rows: List[Dict[str, str]] = []
        for row_values in sh.iter_rows(values_only=True, min_row=2):
            raw: Dict[str, Any] = {}
            for i, v in enumerate(row_values):
                if i >= len(headers):
                    continue
                key = headers[i]
                if not key:
                    continue
                raw[key] = v
            entry = {f: _str(raw.get(f, "")) for f in IMPORT_FIELDS}
            if not any(entry[f] for f in DEDUP_FIELDS):
                continue
            rows.append(entry)
        return rows
    finally:
        wb.close()


def _parse_payload(data: bytes, ext: str) -> List[Dict[str, str]]:
    ext = ext.lower()
    if ext == ".csv":
        return _parse_csv(data)
    if ext == ".json":
        return _parse_json(data)
    if ext == ".xlsx":
        return _parse_xlsx(data)
    raise ValueError("Unsupported file type.")


# ------------------------------------------------------------------------------
# Import/Export routes (DB-based)
# ------------------------------------------------------------------------------
@app.route("/export/csv", methods=["GET"])
@login_required
def export_csv() -> Response:
    db = SessionLocal()
    try:
        rows = db.query(Vocabulary).order_by(Vocabulary.created_at.desc()).all()
        items = [_serialize_vocab(v) for v in rows]
    finally:
        db.close()

    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=EXPORT_FIELDS, extrasaction="ignore", lineterminator="\n")
    w.writeheader()
    for it in items:
        w.writerow({f: it.get(f, "") for f in EXPORT_FIELDS})

    filename = f"vocab_export_{export_timestamp()}.csv"
    resp = Response(buf.getvalue(), mimetype="text/csv; charset=utf-8")
    resp.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


@app.route("/export/html", methods=["GET"])
@login_required
def export_html() -> Response:
    db = SessionLocal()
    try:
        rows = db.query(Vocabulary).order_by(Vocabulary.created_at.desc()).all()
        items = [_serialize_vocab(v) for v in rows]
    finally:
        db.close()

    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%SZ")
    html_output = render_template("export.html", items=items, generated_at=generated_at)
    filename = f"vocab_export_{export_timestamp()}.html"
    resp = Response(html_output, mimetype="text/html; charset=utf-8")
    resp.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


@app.route("/export/json", methods=["GET"])
@login_required
def export_json() -> Response:
    db = SessionLocal()
    try:
        rows = db.query(Vocabulary).order_by(Vocabulary.created_at.desc()).all()
        items = [_serialize_vocab(v) for v in rows]
    finally:
        db.close()

    payload = json.dumps([{f: it.get(f, "") for f in EXPORT_FIELDS} for it in items], ensure_ascii=False, indent=2)
    filename = f"vocab_export_{export_timestamp()}.json"
    resp = Response(payload, mimetype="application/json; charset=utf-8")
    resp.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


@app.route("/export/xlsx", methods=["GET"])
@login_required
def export_xlsx() -> Response:
    db = SessionLocal()
    try:
        rows = db.query(Vocabulary).order_by(Vocabulary.created_at.desc()).all()
        items = [_serialize_vocab(v) for v in rows]
    finally:
        db.close()

    wb = Workbook()
    sh = wb.active
    sh.title = "Vocabulary"

    sh.append(list(EXPORT_FIELDS))
    header_font = Font(bold=True)
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))

    for c in sh[1]:
        c.font = header_font
        c.border = thin

    for it in items:
        row = [it.get(f, "") for f in EXPORT_FIELDS]
        sh.append(row)
        for c in sh[sh.max_row]:
            c.border = thin

    for idx, col in enumerate(sh.columns, start=1):
        maxlen = max((len(str(c.value)) for c in col if c.value is not None), default=0)
        sh.column_dimensions[get_column_letter(idx)].width = min(maxlen + 2, 60) if maxlen else 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"vocab_export_{export_timestamp()}.xlsx"
    resp = Response(buf.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


@app.route("/import", methods=["POST"])
@login_required
def import_words() -> Response:
    file = request.files.get("file")
    if file is None or not file.filename:
        flash("Please choose a file to import.", "error")
        return redirect(url_for("vocabulary"))

    filename = secure_filename(file.filename)
    ext = Path(filename).suffix.lower()
    if ext not in ALLOWED_IMPORT_EXTENSIONS:
        flash("Unsupported file type. Please upload a CSV, JSON, or XLSX file.", "error")
        return redirect(url_for("vocabulary"))

    data = file.read()
    if not data:
        flash("Uploaded file is empty.", "warning")
        return redirect(url_for("vocabulary"))

    try:
        rows = _parse_payload(data, ext)
    except ValueError as exc:
        flash(str(exc), "error")
        return redirect(url_for("vocabulary"))

    if not rows:
        flash("No valid rows found in the uploaded file.", "warning")
        return redirect(url_for("vocabulary"))

    # نگه داشتن برای preview
    session[PENDING_IMPORT_SESSION_KEY] = rows
    session[PENDING_IMPORT_META_SESSION_KEY] = {
        "total_rows": len(rows),
        "duplicate_rows": 0,            # در confirm حساب می‌کنیم
        "skipped_missing_word": 0,      # در confirm حساب می‌کنیم
        "filename": filename,
    }
    session.modified = True
    return redirect(url_for("import_preview"))


@app.route("/import/preview", methods=["GET"])
@login_required
def import_preview() -> Response | str:
    pending_rows = session.get(PENDING_IMPORT_SESSION_KEY)
    meta = session.get(PENDING_IMPORT_META_SESSION_KEY)

    if not pending_rows or not meta:
        flash("No import in progress.", "warning")
        return redirect(url_for("vocabulary"))

    total_rows = int(meta.get("total_rows", len(pending_rows)))
    filename = meta.get("filename", "")
    preview_limit = 50
    preview_rows = pending_rows[:preview_limit]

    return render_template(
        "import_preview.html",
        filename=filename,
        total_rows=total_rows,
        new_count=len(pending_rows),
        duplicate_count=int(meta.get("duplicate_rows", 0)),
        skipped_missing_word=int(meta.get("skipped_missing_word", 0)),
        preview_rows=preview_rows,
        preview_limit=preview_limit,
    )


@app.route("/import/confirm", methods=["POST"])
@login_required
def confirm_import() -> Response:
    pending_rows = session.get(PENDING_IMPORT_SESSION_KEY) or []
    if not pending_rows:
        flash("No rows pending import.", "warning")
        return redirect(url_for("vocabulary"))

    imported = 0
    duplicate = 0
    missing_word = 0

    db = SessionLocal()
    try:
        for row in pending_rows:
            word = _str(row.get("word"))
            if not word:
                missing_word += 1
                continue

            sentence = _str(row.get("sentence"))
            synonym = _str(row.get("synonym"))
            pos = _str(row.get("type"))
            base_word = _str(row.get("base_word"))
            native_meaning = _str(row.get("native_meaning"))
            created_raw = _str(row.get("created_at"))

            exists = (
                db.query(Vocabulary.id)
                .filter(
                    Vocabulary.word == word,
                    Vocabulary.sentence == sentence,
                    Vocabulary.synonym == synonym,
                    Vocabulary.pos == pos,
                    Vocabulary.base_word == base_word,
                    Vocabulary.native_meaning == native_meaning,
                )
                .first()
            )
            if exists:
                duplicate += 1
                continue

            created_at = None
            if created_raw:
                try:
                    created_at = datetime.fromisoformat(created_raw.replace("Z", "+00:00"))
                except Exception:
                    created_at = None
            if created_at is None:
                created_at = datetime.now(timezone.utc)

            db.add(
                Vocabulary(
                    word=word,
                    sentence=sentence,
                    synonym=synonym,
                    pos=pos,
                    base_word=base_word,
                    native_meaning=native_meaning,
                    created_at=created_at,
                )
            )
            imported += 1

        db.commit()
        flash(f"Imported {imported} new {'entry' if imported == 1 else 'entries'}.", "success")
    except Exception as e:
        db.rollback()
        flash(f"Import failed: {e}", "error")
    finally:
        session.pop(PENDING_IMPORT_SESSION_KEY, None)
        session.pop(PENDING_IMPORT_META_SESSION_KEY, None)
        session.modified = True
        db.close()

    return redirect(url_for("vocabulary"))


# ------------------------------------------------------------------------------
# Run
# ------------------------------------------------------------------------------
if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5000, debug=False)
